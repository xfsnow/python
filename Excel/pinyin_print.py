from openpyxl import load_workbook
from pypinyin import pinyin, lazy_pinyin, Style


# t = [1, 2, 3, 1, 2, 5, 6, 7, 8]
# unique = list(set(t))
# print(unique)
# exit()

def array_flatten(array, separator=','):
    e = [y for x in array for y in x]
    string = separator.join(e)
    return string

string = '''看见	哪里	那边	头顶	眼睛
雪白	肚皮	孩子	天空	傍晚
人们	冬天	花朵	平常	江河
海洋	田地	工作	办法	如果
长大	四海为家	娃娃	只要
皮毛	那里	知识	花园	石桥
队旗	铜号	红领巾	欢笑
树叶	枫树	松柏	木棉	水杉
化石	金桂	写字	丛林	深处
竹林	熊猫	朋友	四季	农事
月光	辛苦	棉衣	一同	柱子
到底	秤杆	力气	出来	船身
石头	地方	果然	评奖	时间
报纸	来不及	事情	坏事
出国	好事 	今天	圆珠笔
开心	以前	还有	台灯'''

# string = '看见	哪里	看见	哪里    那边	头顶	眼睛'
phrases = string.split()
# 去重
# phrases = list(set(phrases))
# 按汉字排序，其实是没有顺序的
# phrases.sort()
# print(phrases)

arr = []
for phrase in phrases:
    arr_py = pinyin(phrase, style=Style.TONE)
    py = array_flatten(arr_py, ' ')
    arr.append(py)
folder = './Excel/'
excel_template = folder + 'pinyin_square_block.xlsx'
wb = load_workbook(filename = excel_template)
ws = wb.active
r = 1
c = 1
for i in range(0, len(phrases)):
    phrase = phrases[i]
    arr_py = pinyin(phrase, style=Style.TONE)
    py = array_flatten(arr_py, ' ')
    d = ws.cell(row=r, column=c, value=py)
    # 每行10个汉字。一个汉字 2 个单元格，所以下一个词组开始的位置是
    c = c + len(arr_py) * 2
    # 如果出现 3 字词，则剩下不够 2 字词了，直接换新行。
    if (c > 18):
        # 每行字实际是3 个表格行
        r = r + 3
        c = 1

wb.save(folder + 'pinyin119.xlsx')