from openpyxl import load_workbook

# str = '''看见	哪里	那边	头顶	眼睛
# 雪白	肚皮	孩子	天空	傍晚
# 人们	冬天	花朵	平常	江河
# 海洋	田地	工作	办法	如果
# 长大	四海为家	娃娃	只要
# 皮毛	那里	知识	花园	石桥
# 队旗	铜号	红领巾	欢笑
# 树叶	枫树	松柏	木棉	水杉
# 化石	金桂	写字	丛林	深处
# 竹林	熊猫	朋友	四季	农事
# 月光	辛苦	棉衣	一同	柱子
# 到底	秤杆	力气	出来	船身
# 石头	地方	果然	评奖	时间
# 报纸	来不及	事情	坏事
# 出国	好事 	今天	圆珠笔
# 开心	以前	还有	台灯'''

# phrases = str.split()
# print(phrases)

folder = './Excel/'
# excel_template = folder + 'pinyin_square_block.xlsx'
# wb = load_workbook(filename = excel_template)
# ws = wb['pinyin']
# pageTitle = '语文二年级上册119页汉字'
# ws.title = pageTitle
# d = ws.cell(row=1, column=3, value=pageTitle)
# # print(ws['A1'].value)
# wb.save(folder + 'pinyin119.xlsx')

# 读取当前目录下的 black_dots.xlsx 文件
wb = load_workbook(filename = folder + 'black_dots.xlsx')
# 读取第一个工作表
ws = wb.active
# 读取第 16 行，L 到 R 单元格的值
for row in ws.iter_rows(min_row=16, max_row=16, min_col=12, max_col=18):
    # 输出行号
    print('Row:', row[0].row)
    for cell in row:
        # 输出单元格列号及，单元格的值
        print('Column:', cell.column, 'Value:', cell.value)
        content = cell.value
        # 如果 content 非None并且 Unicode 编码是 b'\xe2\x97\x8f' 输出 1，否则输出 0
        if content and content.encode() == b'\xe2\x97\x8f':
            print(1)
        else:
            print(0)